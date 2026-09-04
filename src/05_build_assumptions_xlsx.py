"""
05_build_assumptions_xlsx.py
Builds reports/assumptions.xlsx: the sourced, editable assumptions register
that drives the cost and stress-threshold sensitivity of the analysis.

Run:  python src/05_build_assumptions_xlsx.py   (after 04_export_for_bi.py)
"""

from pathlib import Path
import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

DB_PATH = Path("data/interim/grid.duckdb")
MANIFEST_PATH = Path("reports/pull_manifest.csv")
QUALITY_PATH = Path("reports/data_quality_summary.csv")
OUT_PATH = Path("reports/assumptions.xlsx")

BLUE = PatternFill("solid", fgColor="DDEBF7")
RED = PatternFill("solid", fgColor="FFC7CE")
BOLD = Font(bold=True)

def main():
    con = duckdb.connect(str(DB_PATH), read_only=True)
    manifest = pd.read_csv(MANIFEST_PATH)
    window_start = manifest["window_start"].iloc[0]
    window_end = manifest["window_end"].iloc[0]

    peak = con.execute("""
        SELECT ba_code, ROUND(MAX(demand_mw)) AS peak_mw
        FROM fact_grid_hourly GROUP BY ba_code ORDER BY peak_mw DESC
    """).df()

    worst_region = con.execute("""
        SELECT ba_code FROM analysis_stress_penalty
        ORDER BY stress_penalty_pp DESC LIMIT 1
    """).fetchone()[0]

    wb = Workbook()

    # --- Tab 1: Assumptions -----------------------------------------------------
    ws = wb.active
    ws.title = "Assumptions"
    headers = ["Parameter", "Value", "Unit", "Source / rationale"]
    ws.append([])
    ws.append(headers)
    for c in range(1, 5):
        ws.cell(row=2, column=c).font = BOLD

    rows = [
        ("Stress hour threshold", 0.95, "percentile", "Top 5% of demand within region and season. Sensitivity run at 0.90 and 0.99."),
        ("Extreme hour threshold", 0.99, "percentile", "Tail case"),
        ("Analysis window start", window_start, "date", "24 months back from pull, from the manifest"),
        ("Analysis window end", window_end, "date", "Start of pull month, from the manifest"),
        ("Imbalance tolerance", 5.0, "percent", "Above this, an hour is flagged in QA"),
        ("Shortfall price, low", 40, "USD per MWh", "EIA Short-Term Energy Outlook, Jan 2025: 2025 U.S. demand-weighted average wholesale price"),
        ("Shortfall price, central", 100, "USD per MWh", "EIA Today in Energy, summer 2022: on-peak wholesale prices near $98-100/MWh in CAISO/Northeast heat events"),
        ("Shortfall price, high", 200, "USD per MWh", "EIA Today in Energy: ERCOT North hub averaged $182/MWh in July 2022 record demand"),
    ]
    for r in rows:
        ws.append(list(r))

    for row in range(3, 11):
        ws.cell(row=row, column=2).fill = BLUE

    ws.cell(row=12, column=1, value="Blue cells are inputs. Everything else is calculated.")
    for col, width in zip("ABCD", [26, 14, 14, 70]):
        ws.column_dimensions[col].width = width

    # --- Tab 2: BA_Reference -----------------------------------------------------
    ws2 = wb.create_sheet("BA_Reference")
    ws2.append(["Code", "Name", "Region", "Local timezone", "Peak demand (MW)"])
    for c in range(1, 6):
        ws2.cell(row=1, column=c).font = BOLD

    names = {
        "PJM": ("PJM Interconnection", "Mid Atlantic", "America/New_York"),
        "MISO": ("Midcontinent ISO", "Midwest", "America/New_York"),
        "CISO": ("California ISO", "West", "America/Los_Angeles"),
        "ERCO": ("ERCOT", "Texas", "America/Chicago"),
        "ISNE": ("ISO New England", "Northeast", "America/New_York"),
        "NYIS": ("New York ISO", "Northeast", "America/New_York"),
        "SWPP": ("Southwest Power Pool", "Central", "America/Chicago"),
        "SOCO": ("Southern Company", "Southeast", "America/New_York"),
    }
    peak_lookup = dict(zip(peak["ba_code"], peak["peak_mw"]))
    for code, (name, region, tz) in names.items():
        ws2.append([code, name, region, tz, peak_lookup.get(code)])
    for col, width in zip("ABCDE", [8, 24, 14, 20, 18]):
        ws2.column_dimensions[col].width = width

    # --- Tab 3: Data_Quality -----------------------------------------------------
    ws3 = wb.create_sheet("Data_Quality")
    qa = pd.read_csv(QUALITY_PATH)
    ws3.append(list(qa.columns))
    for c in range(1, len(qa.columns) + 1):
        ws3.cell(row=1, column=c).font = BOLD
    for _, r in qa.iterrows():
        ws3.append(list(r))

    pct_complete_col = qa.columns.get_loc("pct_complete") + 1
    imbalance_col = qa.columns.get_loc("median_abs_imbalance_pct") + 1
    n = len(qa)
    pc_letter = get_column_letter(pct_complete_col)
    im_letter = get_column_letter(imbalance_col)
    ws3.conditional_formatting.add(
        f"{pc_letter}2:{pc_letter}{n+1}",
        CellIsRule(operator="lessThan", formula=["99"], fill=RED),
    )
    ws3.conditional_formatting.add(
        f"{im_letter}2:{im_letter}{n+1}",
        CellIsRule(operator="greaterThan", formula=["1"], fill=RED),
    )
    for i, col in enumerate(qa.columns, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

    # --- Tab 4: Sensitivity -------------------------------------------------------
    ws4 = wb.create_sheet("Sensitivity")
    ws4.cell(row=1, column=1, value=f"Stress shortfall UPPER BOUND ($M), worst region ({worst_region})").font = BOLD

    costs = [40, 100, 200]
    thresholds = [0.90, 0.95, 0.99]
    ws4.append([])
    ws4.append(["", "$40/MWh", "$100/MWh", "$200/MWh"])
    for c in range(1, 5):
        ws4.cell(row=2, column=c).font = BOLD

    for th in thresholds:
        shortfall = con.execute("""
            SELECT ROUND(SUM(CASE WHEN demand_pctile >= ? THEN shortfall_mw ELSE 0 END))
            FROM fact_grid_hourly WHERE ba_code = ?
        """, [th, worst_region]).fetchone()[0] or 0
        row_vals = [th] + [round(shortfall * cost / 1_000_000, 2) for cost in costs]
        ws4.append(row_vals)

    for col, width in zip("ABCD", [10, 12, 12, 12]):
        ws4.column_dimensions[col].width = width

    con.close()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
